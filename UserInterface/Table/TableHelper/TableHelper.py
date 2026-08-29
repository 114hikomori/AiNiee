import openpyxl
import rapidjson as json
from qfluentwidgets import SwitchButton, TableWidget
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QHBoxLayout, QTableWidgetItem, QWidget

class TableHelper():

    def __init__(self) -> None:
        super().__init__()

    # 从表格加载数据
    def load_from_table(table: TableWidget, keys: list[str], switch_keys=None) -> list[dict]:
        switch_keys = set(switch_keys or ())
        result = []

        # 遍历每一行
        for row in range(table.rowCount()):
            # 获取当前行所有条目
            data: list[QTableWidgetItem] = [table.item(row, col) for col in range(table.columnCount())]

            # 检查数据合法性
            if not isinstance(data[0], QTableWidgetItem) or len(data[0].text().strip()) == 0:
                continue

            # 添加数据
            row_data = {}
            for i, key in enumerate(keys):
                if key in switch_keys:
                    switch = TableHelper.get_switch_from_cell(table, row, i)
                    row_data[key] = switch.isChecked() if isinstance(switch, SwitchButton) else False
                else:
                    row_data[key] = data[i].text().strip() if isinstance(data[i], QTableWidgetItem) else ""
            result.append(row_data)

        return result

    # 向表格更新数据
    def update_to_table(table: TableWidget, data: list[dict], keys: list[str], switch_keys=None, switch_changed=None) -> None:
        switch_keys = set(switch_keys or ())

        # 清理旧的开关容器，避免刷新表格后留下由 Qt 托管的隐藏控件。
        switch_columns = [i for i, key in enumerate(keys) if key in switch_keys]
        for row in range(table.rowCount()):
            for col in switch_columns:
                container = table.cellWidget(row, col)
                if container is not None:
                    table.removeCellWidget(row, col)
                    container.deleteLater()

        # 去重并丢弃无首列内容的空条目。
        data_unique = {}
        for value in data:
            source = str(value.get(keys[0], "") or "").strip()
            if source:
                data_unique[source] = value
        data = list(data_unique.values())

        # 设置表格行数
        row_count = max(16, len(data))
        table.clearContents()
        table.setRowCount(row_count)

        # 遍历表格；正则开关只在首列有内容时显示。
        for row in range(row_count):
            v = data[row] if row < len(data) else {}
            for col in range(table.columnCount()):
                key = keys[col]
                if key in switch_keys:
                    if row < len(data):
                        TableHelper.set_switch_cell(
                            table,
                            row,
                            col,
                            v.get(key) is True,
                            switch_changed,
                        )
                    else:
                        TableHelper.clear_switch_cell(table, row, col)
                    continue

                value = v.get(key, "")
                table.setItem(row, col, QTableWidgetItem("" if value is None else str(value)))

    @staticmethod
    def get_switch_from_cell(table: TableWidget, row: int, col: int):
        container = table.cellWidget(row, col)
        if isinstance(container, SwitchButton):
            return container
        return container.findChild(SwitchButton) if isinstance(container, QWidget) else None

    @staticmethod
    def clear_switch_cell(table: TableWidget, row: int, col: int) -> None:
        container = table.cellWidget(row, col)
        if container is not None:
            table.removeCellWidget(row, col)
            container.deleteLater()

        item = QTableWidgetItem("")
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        table.setItem(row, col, item)

    @staticmethod
    def sync_switch_cell(table: TableWidget, row: int, source_col: int, switch_col: int, switch_changed=None):
        source_item = table.item(row, source_col)
        has_source = isinstance(source_item, QTableWidgetItem) and bool(source_item.text().strip())
        switch = TableHelper.get_switch_from_cell(table, row, switch_col)

        if has_source:
            if not isinstance(switch, SwitchButton):
                switch = TableHelper.set_switch_cell(
                    table,
                    row,
                    switch_col,
                    False,
                    switch_changed,
                )
            return switch

        TableHelper.clear_switch_cell(table, row, switch_col)
        return None

    @staticmethod
    def set_switch_cell(table: TableWidget, row: int, col: int, checked: bool = False, switch_changed=None):
        old_container = table.cellWidget(row, col)
        if old_container is not None:
            table.removeCellWidget(row, col)
            old_container.deleteLater()

        item = QTableWidgetItem("")
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        table.setItem(row, col, item)

        container = QWidget(table)
        layout = QHBoxLayout(container)
        layout.setContentsMargins(4, 0, 4, 0)
        layout.setAlignment(Qt.AlignCenter)

        switch = SwitchButton(container)
        switch.setOnText("")
        switch.setOffText("")
        switch.setChecked(bool(checked))
        if switch_changed is not None:
            switch.checkedChanged.connect(switch_changed)
        layout.addWidget(switch)
        table.setCellWidget(row, col, container)
        return switch

    # 从文件加载数据
    def load_from_file(path: str, keys: list[str], switch_keys=None) -> list[dict]:
        result = []

        # 从 json 文件加载数据
        if path.endswith(".json"):
            result = TableHelper.load_from_json_file(path, keys, switch_keys)

        # 从 xlsx 文件加载数据
        if path.endswith(".xlsx"):
            result = TableHelper.load_from_xlsx_file(path, keys, switch_keys)

        return result

    # 从 json 文件加载数据
    def load_from_json_file(path: str, keys: list[str], switch_keys=None) -> list[dict]:
            switch_keys = set(switch_keys or ())
            result = []

            # 读取文件
            inputs = []
            with open(path, "r", encoding = "utf-8") as reader:
                inputs = json.load(reader)

            # 标准字典列表
            # [
            #     {
            #         "key": "value",
            #         "key": "value",
            #         "key": "value",
            #     }
            # ]
            if isinstance(inputs, list):
                for data in inputs:
                    # 数据校验
                    if not isinstance(data, dict) or str(data.get(keys[0], "")).strip() == "":
                        continue

                    # 添加数据
                    row = {}
                    for key in keys:
                        value = data.get(key, "")
                        if key in switch_keys and isinstance(value, bool):
                            row[key] = value
                        else:
                            row[key] = str(value).strip() if value is not None else ""
                    result.append(row)

                # 兼容旧版，保留一段时间用以过度
                if len(result) == 0 and "src" in keys:
                    # 将 keys 中的 "src" 替换为 "srt" 然后重试
                    result = TableHelper.load_from_json_file(
                        path,
                        [("srt" if v == "src" else v) for v in keys],
                        switch_keys,
                    )

                    # 将字段换回来
                    for v in result:
                        v["src"] = v.get("srt", "")

            # 标准 KV 字典
            # [
            #     "ダリヤ": "达莉雅"
            # ]
            if isinstance(inputs, dict):
                for k, v in inputs.items():
                    # 数据校验
                    if str(k).strip() == "":
                        continue

                    # 添加数据
                    item = {}
                    for i in range(len(keys)):
                        if i == 0:
                            item[keys[i]] = str(k).strip()
                        elif i == 1:
                            item[keys[i]] = str(v).strip() if v != None else ""
                        else:
                            item[keys[i]] = False if keys[i] in switch_keys else ""
                    result.append(item)

            return result

    # 从 xlsx 文件加载数据
    def load_from_xlsx_file(path: str, keys: list[str], switch_keys=None) -> list[dict]:
        switch_keys = set(switch_keys or ())
        result = []

        sheet = openpyxl.load_workbook(path).active
        for row in range(2, sheet.max_row + 1): # 跳过标题行，从第二行开始
            # 读取每一行的数据
            data: list[str] = [
                sheet.cell(row = row, column = col).value
                for col in range(1, len(keys) + 1)
            ]

            # 检查数据合法性
            if data[0] == None or str(data[0]).strip() == "":
                continue

            # 添加数据
            row_data = {}
            for i, key in enumerate(keys):
                value = data[i]
                if key in switch_keys and isinstance(value, bool):
                    row_data[key] = value
                else:
                    row_data[key] = str(value).strip() if value is not None else ""
            result.append(row_data)

        return result
